# -*- coding: utf-8 -*-
""" 
@Author:MaoMorn
@Time: 2018/08/30
@Description: excel的相关操作
"""
import openpyxl
import re
import os


def write07excel(path):
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = '2007测试表'

    value = [["名称", "价格", "出版社", "语言"],
             ["如何高效读懂一本书", "22.3", "机械工业出版社", "中文"],
             ["暗时间", "32.4", "人民邮电出版社", "中文"],
             ["拆掉思维里的墙", "26.7", "机械工业出版社", "中文"]]
    for i in range(0, 4):
        for j in range(0, len(value[i])):
            sheet.cell(row=i + 1, column=j + 1, value=str(value[i][j]))

    wb.save(path)
    print("写入数据成功！")


def read07excel(path):
    wb = openpyxl.load_workbook(path)
    sheet = wb['Sheet1']
    result = []
    for row in sheet["B"]:
        if type(row.value) == str:
            match = re.compile("/(((?!/)\S)+)\.git$").search(row.value)
            if match:
                pro_name = match.group(1)
                if pro_name == "android":
                    pro_name = re.compile("/(((?!/)\S)+)/android\.git$").search(row.value).group(1)
                result.append([pro_name, row.value])
                # print(match.group(1))
    return result


def git_clone(dir_path, url):
    cmd = 'git clone %s "%s"' % (url, dir_path)
    os.system(cmd)


if __name__ == "__main__":
    excel_path = "E:\\硕士研究\\SoftwareTest\\开源应用列表.xlsx"
    result = read07excel(excel_path)
    target_dir = "E:\\硕士研究\\SoftwareTest\\测试应用\测试应用源码"
    for i in range(len(result)):
        j = i + 1
        while j < len(result):
            if result[i][0] == result[j][0]:
                print(result[i][1])
                print(result[j][1])
            j += 1
    # for item in result:
    #     dir_path = os.path.join(target_dir, item[0])
    #     url = item[1]
    #     git_clone(dir_path, url)
    # pass
    # test = "https://github.com/MalaysiaPrayerTimes/android.git"
