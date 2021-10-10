# -*- coding: utf-8 -*-
""" 
@Author:MaoMorn
@Time: 2020/06/19
@Description: 
"""
import openpyxl
import csv
import os
import scipy.stats
import itertools
import cliffsDelta


def read_by_openpyxl(dir_path, name):
    name += '.xlsx'
    xlsx_file = openpyxl.load_workbook(os.path.join(dir_path, name))
    dir_path = os.path.join(dir_path, name[:name.find('.')])
    if not os.path.exists(dir_path):
        os.mkdir(dir_path)
    sheets = xlsx_file.sheetnames
    # 循环遍历所有sheet
    for i in range(1, len(sheets)):
        sheet = xlsx_file[sheets[i]]
        print('\n\n第' + str(i) + '个sheet: ' + sheet.title + '->>>')
        index = sheet.title.find('_')
        log = sheet.title[:index].lower()
        smell = sheet.title[index + 1:]
        csv_file = open(os.path.join(dir_path, log + '_' + smell + '.csv'), "w", newline='')
        writer = csv.writer(csv_file)
        columns = sheet.columns
        for column in columns:
            # print(column)
            line = [r.value for r in column]
            writer.writerow(line)
        csv_file.close()
    xlsx_file.close()


def pretreatment_csv(directory, name):
    # dir_path = "E:\\Data\\MasterTimes\\论文编写\\重构\\analysis\\soundwaves"
    dir_path = os.path.join(directory, name)
    log_type = ["dumpsys", "gfx", "logcat"]
    smell_type = ["ORIGINAL", "NO_ALL", "NO_HMU", "NO_IGS", "NO_MIM"]
    data = {}
    for log in log_type:
        data[log] = {}
        for smell in smell_type:
            csv_file = open(os.path.join(dir_path, log + "_" + smell + ".csv"), "r")
            reader = csv.reader(csv_file)
            data[log][smell] = []
            for row in reader:
                i = 0
                while i < len(row) and row[i] != '':
                    i += 1
                data[log][smell].append(row[:i])
            csv_file.close()
    for log in log_type:
        for smell in smell_type:
            csv_file = open(os.path.join(dir_path, log + "_" + smell + ".csv"), "w", newline='')
            writer = csv.writer(csv_file)
            for i in data[log][smell]:
                writer.writerow(i)
            csv_file.close()


def save2xslx(data, file_path):
    # 新建文件和表格
    wb = openpyxl.Workbook()
    index = 0
    for log in data:
        sheet = wb.create_sheet(log, index)
        index += 1
        for smell in data[log]:
            row = [smell]
            row.extend(data[log][smell])
            # 添加一行
            sheet.append(row)
    wb.save(file_path)  # 保存文件，注意以xlsx为文件扩展名


def analysis(directory, name):
    dir_path = os.path.join(directory, name)
    log_type = ["dumpsys", "gfx", "logcat"]
    smell_type = ["ORIGINAL", "NO_ALL", "NO_HMU", "NO_IGS", "NO_MIM"]
    data = {}
    # 每组实验数据均值化
    for log in log_type:
        start_index = 1
        if log == 'gfx':
            start_index = 0
        data[log] = {}
        for smell in smell_type:
            csv_file = open(os.path.join(dir_path, log + "_" + smell + ".csv"), "r")
            reader = csv.reader(csv_file)
            data[log][smell] = []
            for row in reader:
                try:
                    data[log][smell].append(sum([eval(i) for i in row[start_index:]]) / (len(row) - 1))
                except:
                    print(log)
                    print(smell)
                    print(row)
            csv_file.close()
    # 显著性计算
    result = {}
    for log in data:
        result[log] = {}
        # print("%s的结果如下：" % log)
        for smell in data[log]:
            # print("%-20s相关信息：" % smell, end='')
            # print(data[log][smell])
            pass
        comb_iter = itertools.combinations(smell_type, 2)
        for comb in comb_iter:
            result[log][comb] = []
            value = scipy.stats.wilcoxon(data[log][comb[0]], data[log][comb[1]])
            result[log][comb].append(value[1])
            value = cliffsDelta.cliffsDelta(data[log][comb[0]], data[log][comb[1]])
            result[log][comb].append(value)
    for log in result:
        for comb in result[log]:
            print(comb)
            print(result[log][comb])
    # for log in result:
    #     print("=" * 30)
    #     print("%-10s的显著性分析统计结果如下：" % log)
    #     print("非常显著(<1%):")
    #     for comb in result[log]:
    #         if result[log][comb] <= 0.01:
    #             print("%-10s和%-10s的pvalue为：%.9f" % (comb[0], comb[1], result[log][comb]))
    #     print("显著(1%~5%):")
    #     for comb in result[log]:
    #         if 0.05 >= result[log][comb] > 0.01:
    #             print("%-10s和%-10s的pvalue为：%.9f" % (comb[0], comb[1], result[log][comb]))
    #     print("无显著(>5%):")
    #     for comb in result[log]:
    #         if result[log][comb] > 0.05:
    #             print("%-10s和%-10s的pvalue为：%.9f" % (comb[0], comb[1], result[log][comb]))

    # file_path = os.path.join(dir_path, name + '.xlsx')
    # save2xslx(data, file_path)
    file_path = os.path.join(dir_path, 'result.xlsx')
    save_result(data, result, file_path)


def save_result(data, result, file_path):
    # 新建文件和表格
    wb = openpyxl.Workbook()
    sheet = wb.create_sheet('result', 0)
    header = ['对象1', '对象2', '对象1均值', '对象2均值', 'pvalue', '显著', 'cliffsDelta', 'cliffs']
    for log in result:
        sheet.append([])
        sheet.append([])
        sheet.append(header)
        for comb in result[log]:
            a, b = comb[0], comb[1]
            avg_a, avg_b = sum(data[log][a]) / len(data[log][a]), sum(data[log][b]) / len(data[log][b])
            tag = '是'
            pvalue = result[log][comb][0]
            cliffs = result[log][comb][1]
            if pvalue > 0.05:
                tag = '否'
            row = [a, b, avg_a, avg_b, pvalue, tag, cliffs[0], cliffs[1]]
            sheet.append(row)
    wb.save(file_path)  # 保存文件，注意以xlsx为文件扩展名


if __name__ == "__main__":
    name = "terminal"
    # name = "soundwaves"
    dir_path = "E:\\Data\\MasterTimes\\论文编写\\重构\\analysis"
    # read_by_openpyxl(dir_path, name)
    pretreatment_csv(dir_path, name)
    analysis(dir_path, name)
    # print(type(eval('58.15')))
    # dir_path="E:\\Data\\MasterTimes\\论文编写\\重构\\analysis\\soundwaves"
    # path = os.path.join(dir_path, name)
    # print(path)
